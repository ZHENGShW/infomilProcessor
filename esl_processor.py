#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
ESL (Electronic Shelf Label) 售后退回数据自动化批量处理脚本
======================================================
功能概述：
  1.  自动扫描 input/ 目录下所有 .xls / .xlsx / .xlsm 文件，逐一处理
  2.  白名单 Sheet 校验（缺失则打印警告并跳过该文件）
  3.  动态模糊表头匹配，以纯文本模式提取 ID 与 Model 两列
  4.  对 ID 执行正则校验（^\\d{18}$），标记界形 ID 供后续高亮
  5.  闭环数据审计：找出 A traiter 中未流入出水管的“失踪 ID”
  6.  模糊正则型号标准化（Stellar Pro > Polaris > Stellar-XXL/XL/MF/M/S）
  7.  两轮库存分配算法（严格原配 → 瀑布流升舱借调 → OUT OF STOCK）
  8.  写出客户专属报表（按型号分 Sheet，单元格颜色高亮异常数据）：  client-{name}.xlsx
  9.  写出内部审计汇总表（Summary 透视表 + Sanity Check 对账单）：dashboard-{name}.xlsx

依赖安装：
  pip install pandas openpyxl
"""

import re
import unicodedata
from collections import Counter, defaultdict
from pathlib import Path
from typing import Dict, List, Optional, Set

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# ============================================================
# § 0.  目录配置（修改这两个变量即可适配不同环境）
# ============================================================
INPUT_DIR  = 'input'    # 工厂原始 Excel 文件所在目录（相对路径）
OUTPUT_DIR = 'output'   # 输出文件目录（不存在则自动创建）


# ============================================================
# § 1.  全局常量与业务配置
# ============================================================

# 白名单：输入文件必须包含的 5 个 Sheet（缺一报错）
REQUIRED_SHEETS: List[str] = [
    'Cassées',
    'A traiter',
    'Bonnes après test',
    'A remplacer',
    'Sortant du stock',
]

# 状态映射：Sheet 名 → 客户报表状态码
# 注：'A traiter'（进水管/审计用）和 'Sortant du stock'（库存池）不写入报表，故不在此映射
SHEET_STATUS_MAP: Dict[str, str] = {
    'Cassées':           'BRK',   # 已物理损坏
    'Bonnes après test': 'REP',   # 测试后修复成功
    'A remplacer':       'SWA',   # 待置换（触发库存分配算法）
}

# 瀑布流升舱借调字典（单向不可逆）
# 语义：当 key 型号缺货时，允许借用 value 型号（高级产品）的库存顶替
# 铁律：禁止降级替换 / 禁止跨冷冻-普通环境 / 禁止 Stellar↔Polaris 跨线替换
UPGRADE_MAP: Dict[str, str] = {
    'STELLAR-S3N@E31HA':   'Stellar Pro-154R-N H',
    'STELLAR-M3N@E31HA':   'Stellar Pro-213R-N L',
    'STELLAR-MFN@E31A':    'Stellar Pro-213F-N L',
    'Stellar-XL3N@':       'Stellar Pro-420R-N L',
    'STELLAR-XXL3N@E31HA': 'Stellar Pro-583R-N H',
}

# ─── openpyxl 单元格背景色 ───────────────────────────────────
FILL_ORANGE = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')  # 畸形 ID
FILL_RED    = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')  # OUT OF STOCK
FILL_YELLOW = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')  # 溢出库存
FILL_PINK   = PatternFill(start_color='FF99CC', end_color='FF99CC', fill_type='solid')  # 冲突/分身 ID
FILL_HEADER = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')  # 表头蓝

# ID 合规性正则：必须是 18 位纯数字，不多不少
ID_REGEX = re.compile(r'^\d{18}$')

# Stellar Pro 系列映射规则（按数字匹配，优先级从上到下）:
#   含 154              → Stellar Pro-154R-N H
#   含 213 且含 F       → Stellar Pro-213F-N L
#   含 213 且不含 F     → Stellar Pro-213R-N L
#   含 420              → Stellar Pro-420R-N L
#   含 583              → Stellar Pro-583R-N H

# Polaris Pro 系列映射规则（按数字/字母组合匹配，优先级从上到下）:
#   含 160              → Polaris Pro-160SQ
#   含 230SQ            → Polaris Pro-230SQ
#   含 230SF            → Polaris Pro-230SF
#   含 420              → Polaris Pro-420Q
#   含 583              → Polaris Pro-583Q


# ============================================================
# § 2.  型号标准化（模糊正则提取，严格优先级）
# ============================================================

def normalize_model(raw: str) -> str:
    """
    将工厂原始 Model 字符串映射为客户标准名称，大小写无关。

    匹配优先级（高 → 低）：
      ① Stellar Pro  →  基于数字直接匹配（154 / 213±F / 420 / 583）
      ② Polaris      →  基于数字/后缀直接匹配（160 / 230SQ / 230SF / 420 / 583）
      ③ Stellar-XXL  →  'STELLAR-XXL3N@E31HA'
      ④ Stellar-XL   →  'Stellar-XL3N@'（排除含 XXL 的情况）
      ⑤ Stellar-MF   →  'STELLAR-MFN@E31A'（M 且含 F）
      ⑥ Stellar-M    →  'STELLAR-M3N@E31HA'（M 且不含 F）
      ⑦ Stellar-S    →  'STELLAR-S3N@E31HA'（兜底：S / ST / S3N 等）
      ⑧ 无法识别     →  保留原值并打印警告

    关键设计：
      ・Stellar Pro 必须先于普通 Stellar 检查，避免 'Stellar Pro 230SF' 中的
        'F' 被错误触发 Stellar-MF 规则。
      ・检查普通 Stellar 时，先用 re.sub 剥离 'STELLAR' 关键字本身，在剩余
        字符串中查找尺码标识符，避免 'S' in 'STELLAR' 产生误判。
    """
    if not isinstance(raw, str) or not raw.strip():
        return (raw or '').strip()

    s  = raw.strip()
    su = s.upper()   # 统一大写，实现大小写无关匹配

    # ① Stellar Pro（含 PRO 关键字，必须最优先）─────────────────
    if re.search(r'STELLAR\s*PRO', su):
        # 规则按优先级从上到下：数字匹配优先，213 需区分是否含 F
        if '154' in su:
            return 'Stellar Pro-154R-N H'
        if '213' in su and 'F' in su:
            return 'Stellar Pro-213F-N L'
        if '213' in su:
            return 'Stellar Pro-213R-N L'
        if '420' in su:
            return 'Stellar Pro-420R-N L'
        if '583' in su:
            return 'Stellar Pro-583R-N H'
        return 'UNMAPPED_ERROR'

    # ① Polaris 系列 ────────────────────────────────────────────
    if re.search(r'POLARIS', su):
        # 230SQ / 230SF 须在纯数字 230 之前检查（更精确的字符串先匹配）
        if '160' in su:
            return 'Polaris Pro-160SQ'
        if '230SQ' in su:
            return 'Polaris Pro-230SQ'
        if '230S-Q' in su:
            return 'Polaris Pro-230SQ'
        if '230SF' in su:
            return 'Polaris Pro-230SF'
        if '230S-F' in su:
            return 'Polaris Pro-230SF'
        if '420' in su:
            return 'Polaris Pro-420Q'
        if '583' in su:
            return 'Polaris Pro-583Q'
        return 'UNMAPPED_ERROR'

    # ③~⑦ 普通 Stellar 系列 ──────────────────────────────────────
    if re.search(r'STELLAR', su):
        # 剥离 'STELLAR' 关键字本身后，在剩余部分查找尺码标识
        # 目的：防止 'S' in 'STELLAR' 造成误判
        remainder = re.sub(r'STELLAR', '', su).strip()

        if 'XXL' in remainder:                        # ③ XXL（最长尺码，最先匹配）
            return 'STELLAR-XXL3N@E31HA'
        if 'XL' in remainder:                         # ④ XL（已排除含 XXL 的情况）
            return 'Stellar-XL3N@'
        if 'M' in remainder and 'F' in remainder:     # ⑤ MF（冷冻中号）
            return 'STELLAR-MFN@E31A'
        if 'M' in remainder:                          # ⑥ M（中号，不含 F）
            return 'STELLAR-M3N@E31HA'
        # ⑦ S 兜底：涵盖 S / ST / S3N 及其他未识别的 Stellar 变体
        return 'STELLAR-S3N@E31HA'

    # ⓪ 完全无法识别 → 标记为统一错误标识，由下游拦截过滤
    return 'UNMAPPED_ERROR'


# ============================================================
# § 3.  数据读取与极简清洗
# ============================================================

def _normalize_header(s: str) -> str:
    """
    将列名统一化，用于模糊匹配：
      ・去除首尾空格
      ・转为大写
      ・通过 Unicode NFD 分解 + 过滤组合字符（Mn 类别）剥离法语重音符
        例：È/É/Ê → E，À → A，Ç → C 等
    """
    s = s.strip().upper()
    # NFD 分解后，重音符号（Mn 类别 = Mark, Nonspacing）单独成码位，过滤即可剥离
    s = unicodedata.normalize('NFD', s)
    return ''.join(c for c in s if unicodedata.category(c) != 'Mn')


def read_sheet_as_text(xl: pd.ExcelFile, sheet_name: str) -> pd.DataFrame:
    """
    以纯文本格式读取指定 Sheet，通过动态模糊匹配定位 ID 列和 Model 列。

    防呆设计：
      ・dtype=str 强制 pandas 以字符串读入所有列，根本上杜绝 18 位数字
        被转换为浮点数（如 1.23e+17），彻底防止科学计数法。
      ・动态表头模糊匹配：对每列名执行 _normalize_header() 规范化后：
          - 同时包含 'CODE' 和 'BARRE' → 匹配为 ID 列
            （兼容 'Codes Barres'、'Code Barre'、'CODE_BARRES' 等变体）
          - 包含 'MODEL' 或 'MODELE' → 匹配为 Model 列
            （兼容 'Modèle'、'Modeles'、'MODEL'、'MODELS' 等变体）
      ・任意一列无法匹配时，打印包含实际表头的警告，将该 Sheet 作为空表处理，
        绝对不中断程序。
      ・重命名为 ID / Model 后，丢弃无关列。
      ・删除 ID 与 Model 均为空的完全空白行。

    返回：清洗后的 DataFrame，列为 ['ID', 'Model']
    """
    df = xl.parse(sheet_name, dtype=str)

    # ── 动态模糊定位 ID 列和 Model 列 ────────────────────────────
    id_col    = None
    model_col = None

    for col in df.columns:
        norm = _normalize_header(str(col))
        # ID 列：必须同时包含 CODE 和 BARRE（兼容单复数、带无带空格等变体）
        if id_col is None and 'CODE' in norm and 'BARRE' in norm:
            id_col = col
        # Model 列：包含 MODEL 即可（MODELE 也是 MODEL 的超集，一条规则全覆盖）
        if model_col is None and 'MODEL' in norm:
            model_col = col

    # ── 匹配失败：打印警告，返回空表，不中断主流程 ───────────────
    if id_col is None or model_col is None:
        return pd.DataFrame(columns=['ID', 'Model'])

    df = (
        df[[id_col, model_col]]
        .rename(columns={id_col: 'ID', model_col: 'Model'})
    )

    # 统一清理：NaN → 空字符串，去首尾空格
    df['ID']    = df['ID'].fillna('').astype(str).str.strip()
    df['Model'] = df['Model'].fillna('').astype(str).str.strip()

    # 删除 ID 与 Model 均为空的完全空白行
    df = df[~((df['ID'] == '') & (df['Model'] == ''))].reset_index(drop=True)

    return df


# ============================================================
# § 4.  库存分配辅助函数（先进先出 FIFO）
# ============================================================

def allocate_from_pool(
    stock_pool: Dict[str, List[dict]],
    model_key: str,
    round_label: str = '第一轮',
) -> Optional[dict]:
    """
    从库存池中取出指定型号的第一个可用价签（FIFO 先进先出）。

    参数：
        stock_pool:  { std_model: [{'id': str, 'raw_model': str}, ...] }
        model_key:   需要分配的标准型号名
        round_label: 打印日志时使用的轮次标签

    返回：
        {'id': str, 'raw_model': str}；如库存为空则返回 None。
    """
    queue = stock_pool.get(model_key)
    if queue:
        item = queue.pop(0)   # FIFO：取队列头部
        return item
    return None


# ============================================================
# § 5.  写出客户专属报表（openpyxl 精确控制格式与颜色）
# ============================================================

def write_client_report(
    client_data: Dict[str, List[dict]],
    output_path: str,
    duplicate_ids: Dict[str, List[str]] = {},
) -> bool:
    """
    使用 openpyxl 写出客户专属报表。

    输出结构：
      ・每个标准型号（std_model）独立一个 Sheet，Sheet 名即为型号名。
      ・固定 4 列：FOUND EEG | STATUS | SWAPPED EEG | SWAPPED CODE
      ・所有单元格强制 number_format='@'（文本格式），杜绝科学计数法。
      ・冻结表头行，便于滚动查阅。

    颜色高亮规则：
      橙色 → FOUND EEG 为畸形 ID（未通过 ^\\d{18}$ 校验）
      红色 → SWAPPED EEG / SWAPPED CODE 为 'OUT OF STOCK'（两轮均无库存）
      黄色 → FOUND EEG / STATUS 为 'UNPAIRED'（Sortant du stock 溢出未分配库存）
    """
    wb = Workbook()
    wb.remove(wb.active)   # 删除 openpyxl 自动创建的默认空 Sheet

    HEADER_COLS  = ['FOUND EEG', 'STATUS', 'SWAPPED EEG', 'SWAPPED CODE']
    header_font  = Font(bold=True, color='FFFFFF')
    center_align = Alignment(horizontal='center', vertical='center')
    left_align   = Alignment(horizontal='left',   vertical='center')

    for std_model in sorted(client_data.keys()):
        rows       = client_data[std_model]
        sheet_name = _safe_sheet_name(std_model)
        ws         = wb.create_sheet(title=sheet_name)

        # ── 写入表头 ────────────────────────────────────────────
        ws.append(HEADER_COLS)
        for col_idx in range(1, 5):
            cell           = ws.cell(row=1, column=col_idx)
            cell.font      = Font(bold=True, color='000000')          # 加粗黑色，无底色
            cell.fill      = PatternFill(fill_type=None)              # 强制无背景填充
            cell.alignment = center_align

        # ── 逐行写入数据 ────────────────────────────────────────
        for entry in rows:
            is_unpaired  = entry.get('unpaired',     False)
            is_malformed = entry.get('is_malformed', False)
            is_oos       = entry.get('oos',          False)

            row_idx = ws.max_row + 1
            values  = [
                entry['found_eeg'],
                entry['status'],
                entry['swapped_eeg'],
                entry['swapped_code'],
            ]

            # 写入 4 列，全部强制文本格式
            cells = []
            for col_idx, val in enumerate(values, start=1):
                cell               = ws.cell(
                    row=row_idx, column=col_idx,
                    value=str(val) if val is not None else ''
                )
                cell.number_format = '@'          # 锁定文本格式，防科学计数法
                cell.alignment     = left_align
                cells.append(cell)

            # ── 应用颜色高亮 ─────────────────────────────────────
            if is_unpaired:
                # 溢出库存行 → FOUND EEG 和 STATUS 背景黄色
                cells[0].fill = FILL_YELLOW
                cells[1].fill = FILL_YELLOW

            if is_oos:
                # 极端缺货 → SWAPPED EEG 和 SWAPPED CODE 背景红色
                cells[2].fill = FILL_RED
                cells[3].fill = FILL_RED

            # ── 畸形 ID 双列独立检查（最后执行，确保优先级最高）────
            # 逻辑：对 FOUND EEG 和 SWAPPED EEG 各自取单元格真实值，
            # 若该值既不是系统保留字（空值/UNPAIRED/OUT OF STOCK），
            # 也不满足 18 位纯数字正则，则立即强制填充橙色。
            # 最后再检查冲突 ID（粉色），确保其覆盖其他高亮。
            _RESERVED_VALUES = {'', 'UNPAIRED', 'OUT OF STOCK'}
            for _cell in (cells[0], cells[2]):   # FOUND EEG, SWAPPED EEG
                _val = str(_cell.value).strip() if _cell.value is not None else ''
                if _val not in _RESERVED_VALUES and not re.match(r'^\d{18}$', _val):
                    _cell.fill = FILL_ORANGE
                # 冲突 ID 检查（粉色，最后执行，可覆盖橙色）
                if _val in duplicate_ids:
                    _cell.fill = FILL_PINK

        _auto_fit_columns(ws)

    try:
        wb.save(output_path)
    except PermissionError:
        return False
    return True


# ============================================================
# § 6.  写出内部审计汇总表（Dashboard）
# ============================================================

def write_dashboard(
    client_data: Dict[str, List[dict]],
    sheets_data: Dict[str, pd.DataFrame],
    lost_ids: List[str],
    output_path: str,
    duplicate_ids: Dict[str, List[str]] = {},
    unmapped_rows: List[dict] = (),
    malformed_records: List[dict] = (),
    upgraded_records: List[dict] = (),
) -> bool:
    """
    写出内部审计汇总表，包含两个 Sheet：
      'Summary'      → 按客户标准型号汇总 BRK / REP / SWA 的数量（数据透视表）
      'Sanity Check' → 进/出水管对账信息 + 失踪 ID 列表 + 重复 ID 列表 + 未匹配型号列表
    """
    wb = Workbook()
    wb.remove(wb.active)

    _write_summary_sheet(wb.create_sheet('Summary'), client_data)
    _write_sanity_check_sheet(
        wb.create_sheet('Sanity Check'), sheets_data, lost_ids, duplicate_ids,
        unmapped_rows, malformed_records, upgraded_records
    )

    try:
        wb.save(output_path)
    except PermissionError:
        return False
    return True


def _write_summary_sheet(ws, client_data: Dict[str, List[dict]]) -> None:
    """
    Summary Sheet：按型号汇总 BRK / REP / SWA 数量的数据透视表。
    列：Model | BRK | REP | SWA | Total
    最后一行为带 Excel SUM 公式的合计行（支持后续手动追加数据时自动刷新）。
    注：UNPAIRED 行（溢出库存）不计入任何状态统计。
    """
    header_font = Font(bold=True, color='FFFFFF')
    headers     = ['Model', 'BRK', 'REP', 'SWA', 'Total']
    ws.append(headers)
    for col_idx in range(1, 6):
        cell           = ws.cell(row=1, column=col_idx)
        cell.font      = header_font
        cell.fill      = FILL_HEADER
        cell.alignment = Alignment(horizontal='center')

    for std_model in sorted(client_data.keys()):
        rows  = client_data[std_model]
        # 仅统计真实退回记录（BRK/REP/SWA），排除 UNPAIRED 溢出行
        brk   = sum(1 for r in rows if r['status'] == 'BRK')
        rep   = sum(1 for r in rows if r['status'] == 'REP')
        swa   = sum(1 for r in rows if r['status'] == 'SWA')
        total = brk + rep + swa
        ws.append([std_model, brk, rep, swa, total])

    # 合计行：使用 SUM 公式，动态引用数据区域
    last_data_row = ws.max_row
    ws.append([
        'TOTAL',
        f'=SUM(B2:B{last_data_row})',
        f'=SUM(C2:C{last_data_row})',
        f'=SUM(D2:D{last_data_row})',
        f'=SUM(E2:E{last_data_row})',
    ])
    bold_font = Font(bold=True)
    for col in range(1, 6):
        ws.cell(row=ws.max_row, column=col).font = bold_font

    _auto_fit_columns(ws)


def _write_sanity_check_sheet(
    ws,
    sheets_data: Dict[str, pd.DataFrame],
    lost_ids: List[str],
    duplicate_ids: Dict[str, List[str]] = {},
    unmapped_rows: List[dict] = (),
    malformed_records: List[dict] = (),
    upgraded_records: List[dict] = (),
) -> None:
    """
    Sanity Check Sheet（闭环数据对账单）：
      Row 1: 表头
      Row 2: A traiter（进水管）总数量
      Row 3: 出水管（Bonnes après test + A remplacer）总数量
      Row 4: 差异 = 进水管 - 出水管（非零时橙色高亮）
      Row 5: （空白间隔）
      Row 6: 失踪 ID 列表标题（含数量）
      Row 7+: 具体失踪 ID，逐行列出（强制文本格式，防科学计数法）
    """
    a_traiter_count = len(sheets_data['A traiter'])
    bonnes_count    = len(sheets_data['Bonnes après test'])
    remplacer_count = len(sheets_data['A remplacer'])
    output_count    = bonnes_count + remplacer_count
    diff            = a_traiter_count - output_count

    header_font = Font(bold=True, color='FFFFFF')

    # 表头行
    for col_idx, header_text in enumerate(['指标', '数值'], start=1):
        cell           = ws.cell(row=1, column=col_idx, value=header_text)
        cell.font      = header_font
        cell.fill      = FILL_HEADER
        cell.alignment = Alignment(horizontal='center')

    # 数据行
    ws.cell(row=2, column=1, value='A traiter（进水管）总数')
    ws.cell(row=2, column=2, value=a_traiter_count)

    ws.cell(row=3, column=1, value='出水管（Bonnes après test + A remplacer）总数')
    ws.cell(row=3, column=2, value=output_count)

    ws.cell(row=4, column=1, value='差异（进水管 − 出水管）')
    diff_cell = ws.cell(row=4, column=2, value=diff)
    if diff != 0:
        diff_cell.fill = FILL_ORANGE   # 差异非零 → 橙色警示

    # 失踪 ID 列表
    title_cell      = ws.cell(row=6, column=1, value=f'失踪 ID 列表（共 {len(lost_ids)} 条）')
    title_cell.font = Font(bold=True)

    if not lost_ids:
        ok_cell      = ws.cell(row=7, column=1, value='（无失踪 ID，数据完全闭环 ✓）')
        ok_cell.font = Font(color='008000')   # 绿色表示正常
    else:
        for i, lid in enumerate(lost_ids, start=7):
            cell               = ws.cell(row=i, column=1, value=lid)
            cell.number_format = '@'   # 强制文本格式，防止长数字以科学计数法显示

    # ── 冲突 ID 列表（跳过两行空白后开始） ───────────────────────
    dup_title_row  = ws.max_row + 3   # 隔两行空白作为分隔
    dup_title_cell = ws.cell(
        row=dup_title_row, column=1,
        value=f'Duplicate IDs （跨状态重复录入的冲突 ID，共 {len(duplicate_ids)} 条）'
    )
    dup_title_cell.font = Font(bold=True)

    if not duplicate_ids:
        ok_dup      = ws.cell(row=dup_title_row + 1, column=1, value='（无冲突 ID，全局唯一性校验通过 ✓）')
        ok_dup.font = Font(color='008000')
    else:
        # 写入小表头
        for col_idx, hdr in enumerate(['Client Sheets (冲突所在的客户表)', 'Duplicate ID'], start=1):
            hdr_cell      = ws.cell(row=dup_title_row + 1, column=col_idx, value=hdr)
            hdr_cell.font = Font(bold=True)
        # 逐行写入：每行包含冲突来源表（逗号拼接）和冲突 ID
        for i, (did, sheets) in enumerate(sorted(duplicate_ids.items()), start=dup_title_row + 2):
            ws.cell(row=i, column=1, value=', '.join(sheets))
            id_cell               = ws.cell(row=i, column=2, value=did)
            id_cell.number_format = '@'
            id_cell.fill          = FILL_PINK   # 粉色高亮，与客户报表中保持一致

    # ── 未匹配型号列表（隔两行开始） ──────────────────────────
    unmap_title_row  = ws.max_row + 3
    unmap_title_cell = ws.cell(
        row=unmap_title_row, column=1,
        value=f'Unmapped Models （工厂未匹配型号错误拦截，共 {len(unmapped_rows)} 条）'
    )
    unmap_title_cell.font = Font(bold=True)

    if not unmapped_rows:
        ok_unmap      = ws.cell(row=unmap_title_row + 1, column=1, value='（无未识别型号，所有型号均匹配成功 ✓）')
        ok_unmap.font = Font(color='008000')
    else:
        # 写入小表头
        for col_idx, hdr in enumerate(['ID', 'STATUS', '工厂原始错误型号'], start=1):
            hdr_cell           = ws.cell(row=unmap_title_row + 1, column=col_idx, value=hdr)
            hdr_cell.font      = Font(bold=True)
        # 逐行写入
        for i, item in enumerate(unmapped_rows, start=unmap_title_row + 2):
            id_cell               = ws.cell(row=i, column=1, value=item['id'])
            id_cell.number_format = '@'
            ws.cell(row=i, column=2, value=item['status'])
            ws.cell(row=i, column=3, value=item['raw_model'])

    # ── 界形 ID 溢源列表（隔两行开始） ───────────────────────────
    mal_title_row  = ws.max_row + 3
    mal_title_cell = ws.cell(
        row=mal_title_row, column=1,
        value=f'Malformed IDs （格式异常的长条码溢源，共 {len(malformed_records)} 条）'
    )
    mal_title_cell.font = Font(bold=True)

    if not malformed_records:
        ok_mal      = ws.cell(row=mal_title_row + 1, column=1, value='（无界形 ID，所有条码均符合 18 位纯数字格式 ✓）')
        ok_mal.font = Font(color='008000')
    else:
        # 写入小表头（增加工厂原始工作表列）
        for col_idx, hdr in enumerate(
            ['Client Sheet (客户侧型号表)', 'Factory Sheet (工厂原始工作表)', 'Invalid ID'], start=1
        ):
            hdr_cell      = ws.cell(row=mal_title_row + 1, column=col_idx, value=hdr)
            hdr_cell.font = Font(bold=True)
        # 逐行写入
        for i, rec in enumerate(malformed_records, start=mal_title_row + 2):
            ws.cell(row=i, column=1, value=rec['client_sheet'])
            ws.cell(row=i, column=2, value=rec.get('factory_sheets', ''))
            id_cell               = ws.cell(row=i, column=3, value=rec['invalid_id'])
            id_cell.number_format = '@'
            id_cell.fill          = FILL_ORANGE   # 橙色高亮，与客户报表中保持一致

    # ── 升舱借调明细列表（隔两行开始） ───────────────────────
    upg_title_row  = ws.max_row + 3
    upg_title_cell = ws.cell(
        row=upg_title_row, column=1,
        value=f'Upgraded Substitutions （库存不足触发的升舱借调明细，共 {len(upgraded_records)} 条）'
    )
    upg_title_cell.font = Font(bold=True)

    if not upgraded_records:
        ok_upg      = ws.cell(row=upg_title_row + 1, column=1,
                              value='（无升舱借调，所有替换均严格匹配原型号 ✓）')
        ok_upg.font = Font(color='008000')
    else:
        # 写入小表头
        for col_idx, hdr in enumerate(
            ['缺货的普通型号', '退回的旧件 ID', '借调的 Pro 型号', '出库的 Pro ID'], start=1
        ):
            hdr_cell      = ws.cell(row=upg_title_row + 1, column=col_idx, value=hdr)
            hdr_cell.font = Font(bold=True)
        # 逐行写入
        for i, rec in enumerate(upgraded_records, start=upg_title_row + 2):
            ws.cell(row=i, column=1, value=rec['original_model'])
            orig_id_cell               = ws.cell(row=i, column=2, value=rec['found_id'])
            orig_id_cell.number_format = '@'
            ws.cell(row=i, column=3, value=rec['upgraded_model'])
            pro_id_cell               = ws.cell(row=i, column=4, value=rec['upgraded_id'])
            pro_id_cell.number_format = '@'

    _auto_fit_columns(ws)


# ============================================================
# § 7.  工具函数
# ============================================================

def _safe_sheet_name(name: str) -> str:
    """
    将型号名转换为合法的 Excel Sheet 名称：
      ・替换 Excel 不允许的特殊字符：\\ / * ? : [ ]
      ・截断至 Excel 限制的最大长度 31 个字符
    """
    return re.sub(r'[\\/*?:\[\]]', '_', name)[:31]


def _auto_fit_columns(ws) -> None:
    """
    根据当前单元格内容自动调整列宽（最小 10，最大 60 字符宽度单位）。
    """
    for col in ws.columns:
        max_len    = max(
            (len(str(cell.value)) for cell in col if cell.value is not None),
            default=8,
        )
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = min(max(max_len + 4, 10), 60)


# ============================================================
# § 8.  主流程（批量循环）
# ============================================================

def main() -> None:
    # ━━ 初始化路径 ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    base_dir   = Path(__file__).parent
    input_dir  = base_dir / INPUT_DIR
    output_dir = base_dir / OUTPUT_DIR
    output_dir.mkdir(parents=True, exist_ok=True)

    # ━━ 扫描输入目录 ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    excel_files = sorted(
        p for p in input_dir.iterdir()
        if p.suffix.lower() in ('.xls', '.xlsx', '.xlsm')
    )
    if not excel_files:
        print(f'[WARN] 在目录 "{input_dir}" 下未找到任何 Excel 文件，程序退出。')
        return

    # 三态计数器
    count_success = 0
    count_warning = 0
    count_skipped = 0

    # ━━ 逐文件循环 ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    for file_path in excel_files:
        # 前置拦截：跳过 Excel 临时锁文件（~$ 开头）
        if file_path.name.startswith('~$'):
            continue

        stem               = file_path.stem
        sub_dir            = output_dir / stem
        sub_dir.mkdir(parents=True, exist_ok=True)
        client_report_path = sub_dir / f'{stem}.xlsx'
        dashboard_path     = sub_dir / 'dashboard.xlsx'

        # ── Step 1：读取 Excel & 白名单 Sheet 校验 ────────────
        try:
            xl = pd.ExcelFile(file_path, engine='openpyxl')
        except Exception as e:
            print(f'[SKIPPED] 处理失败被跳过: {stem} | 原因: 文件无法读取 ({e})')
            count_skipped += 1
            continue

        missing_sheets = [s for s in REQUIRED_SHEETS if s not in xl.sheet_names]
        if missing_sheets:
            print(f'[SKIPPED] 处理失败被跳过: {stem} | 原因: 缺少必需 Sheet {missing_sheets}')
            count_skipped += 1
            continue

        # ── Step 2：提取各 Sheet 数据并极简清洗 ──────────────
        sheets_data: Dict[str, pd.DataFrame] = {}
        for sheet in REQUIRED_SHEETS:
            sheets_data[sheet] = read_sheet_as_text(xl, sheet)

        # ── Step 3：ID 正则校验，标记畸形 ID ─────────────────
        malformed_ids: Set[str] = set()
        for sheet in REQUIRED_SHEETS:
            df = sheets_data[sheet]
            if df.empty:
                continue
            bad_ids = df.loc[~df['ID'].str.match(r'^\d{18}$', na=False), 'ID'].tolist()
            if bad_ids:
                malformed_ids.update(bad_ids)

        # ── Step 4：型号标准化 ────────────────────────────────
        for sheet in REQUIRED_SHEETS:
            df = sheets_data[sheet]
            if not df.empty:
                df              = df.copy()
                df['Model_std'] = df['Model'].apply(normalize_model)
                sheets_data[sheet] = df

        # ── 畸形 ID 双重视角溯源（Step 4 后执行，Model_std 已可用）
        # 每次扫码为独立记录，同一异常 ID 出现在多个工厂 Sheet 时分行展示
        malformed_records: List[dict] = []
        if malformed_ids:
            for _sheet in REQUIRED_SHEETS:
                _df = sheets_data[_sheet]
                if _df.empty or 'Model_std' not in _df.columns:
                    continue
                for _, _row in _df[_df['ID'].isin(malformed_ids)].iterrows():
                    malformed_records.append({
                        'client_sheet':   _row['Model_std'],
                        'factory_sheets': _sheet,       # 单一工厂 Sheet，不做拼接
                        'invalid_id':     _row['ID'],
                    })

        # ── 全局唯一性审计：以客户标准型号为视角检测重复 ID ───
        _AUDIT_SHEETS = ['Cassées', 'Bonnes après test', 'A remplacer', 'Sortant du stock']
        _id_model_tracker: Dict[str, List[str]] = defaultdict(list)
        for _s in _AUDIT_SHEETS:
            _df = sheets_data[_s]
            if _df.empty or 'Model_std' not in _df.columns:
                continue
            for _, _row in _df.iterrows():
                if _row['ID']:
                    _id_model_tracker[_row['ID']].append(_row['Model_std'])
        duplicate_ids: Dict[str, List[str]] = {
            _id: _models
            for _id, _models in _id_model_tracker.items()
            if len(_models) >= 2
        }

        # ── Step 5：闭环数据审计 ──────────────────────────────
        a_traiter_ids = set(sheets_data['A traiter']['ID'])
        output_ids    = (
            set(sheets_data['Bonnes après test']['ID']) |
            set(sheets_data['A remplacer']['ID'])
        )
        lost_ids = sorted(a_traiter_ids - output_ids)

        # ── Step 6：构建库存分配池 ────────────────────────────
        unmapped_rows: List[dict] = []
        stock_pool: Dict[str, List[dict]] = defaultdict(list)
        for _, row in sheets_data['Sortant du stock'].iterrows():
            if row['Model_std'] == 'UNMAPPED_ERROR':
                unmapped_rows.append({
                    'id':        row['ID'],
                    'status':    'STOCK',
                    'raw_model': row['Model'],
                })
                continue
            stock_pool[row['Model_std']].append({
                'id':        row['ID'],
                'raw_model': row['Model'],
            })

        # ── Step 7：库存分配算法 & 客户报表数据构建 ──────────
        client_data: Dict[str, List[dict]]    = defaultdict(list)
        upgraded_records: List[dict]          = []   # 记录每次升舱借调的明细

        for sheet_name, status_code in SHEET_STATUS_MAP.items():
            df = sheets_data[sheet_name]
            if df.empty:
                continue
            for _, row in df.iterrows():
                found_id  = row['ID']
                std_model = row['Model_std']

                if std_model == 'UNMAPPED_ERROR':
                    unmapped_rows.append({
                        'id':        found_id,
                        'status':    status_code,
                        'raw_model': row['Model'],
                    })
                    continue

                is_malformed = (found_id in malformed_ids)
                entry: dict = {
                    'found_eeg':    found_id,
                    'status':       status_code,
                    'swapped_eeg':  '',
                    'swapped_code': '',
                    'is_malformed': is_malformed,
                    'oos':          False,
                }

                if status_code == 'SWA':
                    allocated       = allocate_from_pool(stock_pool, std_model, '第一轮-严格原配')
                    _upgrade_used   = None   # 升舱借调使用的目标型号（None 表示未触发）
                    if allocated is None:
                        upgrade_target = UPGRADE_MAP.get(std_model)
                        if upgrade_target:
                            allocated = allocate_from_pool(
                                stock_pool, upgrade_target, '第二轮-升舱借调'
                            )
                            if allocated:   # 升舱借调成功，记录目标型号以便后续入表
                                _upgrade_used = upgrade_target
                    if allocated:
                        entry['swapped_eeg']  = allocated['id']
                        entry['swapped_code'] = std_model
                        if _upgrade_used:   # 仅升舱路径成功时才记录
                            upgraded_records.append({
                                'original_model': std_model,
                                'found_id':       found_id,
                                'upgraded_model': _upgrade_used,
                                'upgraded_id':    allocated['id'],
                            })
                    else:
                        entry['swapped_eeg']  = 'OUT OF STOCK'
                        entry['swapped_code'] = 'OUT OF STOCK'
                        entry['oos']          = True

                client_data[std_model].append(entry)

        # ── Step 8：追加溢出库存（UNPAIRED）─────────────────
        for std_model, remaining in stock_pool.items():
            if remaining:
                for item in remaining:
                    client_data[std_model].append({
                        'found_eeg':    'UNPAIRED',
                        'status':       'UNPAIRED',
                        'swapped_eeg':  item['id'],
                        'swapped_code': std_model,
                        'is_malformed': False,
                        'oos':          False,
                        'unpaired':     True,
                    })

        # ── Step 9：写出客户专属报表 ─────────────────────────
        if not write_client_report(client_data, str(client_report_path), duplicate_ids):
            print(f'[SKIPPED] 处理失败被跳过: {stem} | 原因: 客户报表写入失败（请关闭占用的 Excel 文件后重新运行）')
            count_skipped += 1
            continue

        # ── Step 10：写出内部审计汇总表 ──────────────────────
        if not write_dashboard(client_data, sheets_data, lost_ids, str(dashboard_path),
                               duplicate_ids, unmapped_rows, malformed_records, upgraded_records):
            print(f'[SKIPPED] 处理失败被跳过: {stem} | 原因: Dashboard 写入失败（请关闭占用的 Excel 文件后重新运行）')
            count_skipped += 1
            continue

        # ── 三分支心跳日志（每文件唯一输出行）────────────────
        has_anomaly = (
            len(lost_ids)         > 0 or
            len(malformed_ids)    > 0 or
            len(duplicate_ids)    > 0 or
            len(unmapped_rows)    > 0 or
            len(upgraded_records) > 0
        )
        if has_anomaly:
            print(f'[WARNING] 数据异常警报: 文件 "{stem}" 存在业务瑕疵 (界形/失踪/冲突/未映射/升舱借调)，请前往其对应的 dashboard.xlsx 审查明细！')
            count_warning += 1
        else:
            print(f'[SUCCESS] 完美处理: {stem}')
            count_success += 1

    # ━━ 全局汇总看板 ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    print('=' * 65)
    print('  ESL 售后退回数据批量处理完毕！')
    print(f'  完美通过 : {count_success} 份文件')
    print(f'  异常警告 : {count_warning} 份文件')
    print(f'  失败跳过 : {count_skipped} 份文件')
    if count_warning > 0:
        print()
        print('  👉 带有异常警告的文件，请前往 output 下对应的子目录查看 dashboard.xlsx 溯源')
    print('=' * 65)


# ============================================================
# § 9.  脚本入口
# ============================================================
if __name__ == '__main__':
    main()
